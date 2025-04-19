import json
import os
import re
import uuid
from datetime import datetime, timedelta, time
from telegram import Update, ReplyKeyboardMarkup, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment
import pytz
import aiofiles
import aiohttp
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

# Bot tokeni va admin ID
BOT_TOKEN = "7780199997:AAGnfFQ2rnCbrqHjlubLnsyP9whgdK67XFk"
ADMIN_ID = 673491852

# Fayllar
DATA_FILE = "imp.json"
REPORT_FILE = "report.xlsx"

# Google Sheets sozlamalari
SPREADSHEET_ID = "1PtwVVcGwQEQJ4NNVJpv09z1WUTelKUxTruGywXqLoXQ"
SHEET_NAME = "Bot"
CREDENTIALS_FILE = os.path.join(os.path.dirname(__file__), "credentials.json")

# Tashkent vaqti uchun timezone
TASHKENT_TZ = pytz.timezone("Asia/Tashkent")

def get_sheets_service():
    try:
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        service = build("sheets", "v4", credentials=creds)
        return service
    except Exception as e:
        print(f"Google Sheets xizmati bilan bog‘lanishda xato: {e}")
        return None

def load_data():
    try:
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                updated_katalog = {}
                for katalog_name, katalog_data in data.get("katalog", {}).items():
                    if isinstance(katalog_data, list):
                        updated_katalog[katalog_name] = {
                            "id": str(uuid.uuid4()),
                            "items": [{"nom": item["nom"], "id": item.get("id", str(uuid.uuid4())), "rasm_path": item.get("rasm_path")} for item in katalog_data]
                        }
                    else:
                        updated_katalog[katalog_name] = {
                            "id": katalog_data.get("id", str(uuid.uuid4())),
                            "items": katalog_data.get("items", [])
                        }
                data["katalog"] = updated_katalog
                data["qalinlik"] = sorted([str(q) for q in data.get("qalinlik", [])], key=lambda x: float(x.replace(",", ".")))
                save_data(data)
                return data
    except json.JSONDecodeError:
        print("JSON faylida xato, yangi fayl yaratilmoqda...")
    return {
        "obrabotka": {},
        "katalog": {},
        "qalinlik": [],
        "xodimlar": {},
        "orders": []
    }

def save_data(data):
    try:
        data["qalinlik"] = sorted([str(q) for q in data.get("qalinlik", [])], key=lambda x: float(x.replace(",", ".")))
        sorted_obrabotka = dict(sorted(data["obrabotka"].items()))
        for obrabotka in sorted_obrabotka:
            sorted_obrabotka[obrabotka] = sorted(data["obrabotka"][obrabotka])
        data["obrabotka"] = sorted_obrabotka
        data["katalog"] = dict(sorted(data["katalog"].items()))
        for katalog in data["katalog"]:
            data["katalog"][katalog]["items"] = sorted(data["katalog"][katalog]["items"], key=lambda x: x["nom"])
        
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print(f"Faylga yozishda xato: {e}")

def init_excel():
    if not os.path.exists(REPORT_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ҳисобот"
        
        headers = [
            "Сана", "Смена", "Станок", "Тош номи", "Қалинлик",
            "Обработка", "Размери", "Миқдори", "Хажми(м²)", "Изоҳ",
            "Ҳисобга олинмайди", "Тун кунлик", "Кун кунлик", "Ходим номи"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        gray_fill = PatternFill(start_color="31869B", end_color="31869B", fill_type="solid")
        black_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=3, column=col)
            cell.fill = gray_fill
            cell.border = black_border
        
        column_widths = [130, 76, 95, 250, 60, 80, 120, 120, 95, 80, 80, 80, 80, 80]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width / 7
        
        ws['I4'] = '=L4+M4'
        wb.save(REPORT_FILE)

# Tugmalar
ADMIN_KEYBOARD = [
    ["Ҳисобот"],
    ["+ Обработка", "+ Каталог"],
    ["+ Қалинлик", "Ўчириш"],
    ["+ Ходим"]
]

DELETE_KEYBOARD = [
    ["🗑 Обработка", "🗑 Каталог"],
    ["🗑 Станок", "🗑 Қалинлик"],
    ["🗑 Тош"],
    ["Орқага қайтиш"]
]

XODIM_KEYBOARD = [["Янги қўшиш"]]
KATALOG_KEYBOARD = [["Каталог қўшиш", "Орқага қайтиш"]]
ORDER_KEYBOARD = [["Изоҳсиз", "Орқага қайтиш"]]
STANOK_KEYBOARD = [["", ""], ["Орқага қайтиш"]]
REPORT_KEYBOARD = [["Бугунги", "Муддат оралиғи"], ["Орқага қайтиш"]]

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    data = load_data()

    if user_id == ADMIN_ID:
        reply_markup = ReplyKeyboardMarkup(ADMIN_KEYBOARD, resize_keyboard=True)
        await update.message.reply_text("Салом, Админ!", reply_markup=reply_markup)
    elif str(user_id) in data["xodimlar"]:
        reply_markup = ReplyKeyboardMarkup(XODIM_KEYBOARD, resize_keyboard=True)
        await update.message.reply_text(
            "Салом, Ходим!\n\nПастдаги Янги қўшиш тугмаси орқали ишни бошлашингиз мумкин.\n\nДиққат! Сиз киритаётган маълумотлар ва исмингиз админга кўринади, шунинг учун этиборли бўлинг. Малумотларни тўғри киритинг.\n\n/rasm буйруғи орқали каталоглар ва уларнинг ичидаги тошлар билан танишиб чиқишингиз мумкин",
            reply_markup=reply_markup
        )
    else:
        await update.message.reply_text("Кечирасиз, сиз ходим эмассиз!", reply_markup=ReplyKeyboardRemove())

async def rasm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    data = load_data()
    if str(user_id) not in data["xodimlar"] and user_id != ADMIN_ID:
        await update.message.reply_text("Кечирасиз, сиз ходим эмассиз!")
        return

    if not data["katalog"]:
        await update.message.reply_text("Ҳозирча каталоглар йўқ.", reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True))
        return

    keyboard = [[InlineKeyboardButton(name, callback_data=f"rasm_katalog_{info['id']}")] for name, info in sorted(data["katalog"].items())]
    reply_markup = InlineKeyboardMarkup(keyboard)
    context.user_data["last_message_id"] = (await update.message.reply_text("Каталоглардан бирини танланг:", reply_markup=reply_markup)).message_id
    context.user_data["previous_menu"] = None

async def rasmsiz(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    data = load_data()

    if user_id != ADMIN_ID:
        await update.message.reply_text("Кечирасиз, бу фақат админ учун!")
        return

    if context.user_data.get("waiting_for_rasm"):
        katalog = context.user_data["current_katalog"]
        mahsulot = context.user_data["current_mahsulot"]
        mahsulot_id = str(uuid.uuid4())
        if katalog not in data["katalog"]:
            data["katalog"][katalog] = {"id": str(uuid.uuid4()), "items": []}
        data["katalog"][katalog]["items"].append({"nom": mahsulot, "id": mahsulot_id})
        save_data(data)
        context.user_data["waiting_for_rasm"] = False
        context.user_data["current_katalog"] = None
        context.user_data["current_mahsulot"] = None
        await update.message.reply_text(
            f"Тош '{mahsulot}' расмсиз сақланди!",
            reply_markup=ReplyKeyboardMarkup(KATALOG_KEYBOARD, resize_keyboard=True)
        )

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    text = update.message.text
    data = load_data()

    if user_id == ADMIN_ID:
        await handle_admin_message(update, context, data, text)
    elif str(user_id) in data["xodimlar"]:
        await handle_xodim_message(update, context, data, text, user_id)
    else:
        await update.message.reply_text("Кечирасиз, сиз ходим эмассиз!")

async def handle_admin_message(update: Update, context: ContextTypes.DEFAULT_TYPE, data, text):
    if text == "Орқага қайтиш":
        if context.user_data.get("waiting_for_katalog") or context.user_data.get("previous_menu") == "katalog":
            context.user_data["waiting_for_katalog"] = False
            context.user_data["previous_menu"] = None
            context.user_data.clear()
            await update.message.reply_text(
                "Админ панели",
                reply_markup=ReplyKeyboardMarkup(ADMIN_KEYBOARD, resize_keyboard=True)
            )
            return
        elif context.user_data.get("waiting_for_report"):
            context.user_data["waiting_for_report"] = False
            context.user_data.clear()
            await update.message.reply_text(
                "Админ панели",
                reply_markup=ReplyKeyboardMarkup(ADMIN_KEYBOARD, resize_keyboard=True)
            )
            return
        elif context.user_data.get("waiting_for_delete"):
            context.user_data["waiting_for_delete"] = False
            context.user_data.clear()
            await update.message.reply_text(
                "Админ панели",
                reply_markup=ReplyKeyboardMarkup(ADMIN_KEYBOARD, resize_keyboard=True)
            )
            return
        elif context.user_data.get("waiting_for_start_date") or context.user_data.get("waiting_for_end_date"):
            context.user_data["waiting_for_start_date"] = False
            context.user_data["waiting_for_end_date"] = False
            context.user_data["waiting_for_report"] = True
            await update.message.reply_text(
                "Ҳисобот турини танланг:",
                reply_markup=ReplyKeyboardMarkup(REPORT_KEYBOARD, resize_keyboard=True)
            )
            return
        elif context.user_data.get("waiting_for_stanok"):
            context.user_data["waiting_for_stanok"] = False
            context.user_data["current_obrabotka"] = None
            await update.message.reply_text(
                "Админ панели",
                reply_markup=ReplyKeyboardMarkup(ADMIN_KEYBOARD, resize_keyboard=True)
            )
            return
        context.user_data.clear()
        await update.message.reply_text(
            "Админ панели",
            reply_markup=ReplyKeyboardMarkup(ADMIN_KEYBOARD, resize_keyboard=True)
        )
        return

    if context.user_data.get("waiting_for_start_date"):
        await process_date(update, context)
        return

    if context.user_data.get("waiting_for_end_date"):
        await process_date(update, context)
        return

    if context.user_data.get("waiting_for_obrabotka"):
        data["obrabotka"][text] = []
        save_data(data)
        context.user_data["waiting_for_obrabotka"] = False
        context.user_data["current_obrabotka"] = text
        context.user_data["waiting_for_stanok"] = True
        await update.message.reply_text(
            "Обработка учун станок номини киритинг:",
            reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
        )
        return

    if context.user_data.get("waiting_for_stanok"):
        obrabotka = context.user_data.get("current_obrabotka")
        if not obrabotka or obrabotka not in data["obrabotka"]:
            await update.message.reply_text(
                "Обработка топилмади! Илтимос, қайта уриниб кўринг.",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
            context.user_data["waiting_for_stanok"] = False
            context.user_data["current_obrabotka"] = None
            return
        if text not in data["obrabotka"][obrabotka]:
            data["obrabotka"][obrabotka].append(text)
            save_data(data)
            await update.message.reply_text(
                f"Станок '{text}' обработка '{obrabotka}' га муваффақиятли қўшилди. Яна станок қўшасизми?",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
        else:
            await update.message.reply_text(
                f"Станок '{text}' аллақачон '{obrabotka}' обработкасида мавжуд. Яна станок қўшасизми?",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
        return

    if context.user_data.get("waiting_for_katalog"):
        if text in data["katalog"]:
            await update.message.reply_text(
                f"Каталог '{text}' аллақачон мавжуд! Илтимос, бошқа ном киритинг:",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
            return
        katalog_id = str(uuid.uuid4())
        data["katalog"][text] = {"id": katalog_id, "items": []}
        save_data(data)
        context.user_data["waiting_for_katalog"] = False
        context.user_data["current_katalog"] = text
        context.user_data["waiting_for_mahsulot"] = True
        await update.message.reply_text(
            "Тош номини киритинг:",
            reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
        )
        return

    if context.user_data.get("waiting_for_mahsulot"):
        katalog = context.user_data["current_katalog"]
        if not katalog or katalog not in data["katalog"]:
            await update.message.reply_text(
                "Каталог топилмади! Илтимос, қайта уриниб кўринг.",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
            context.user_data.clear()
            return
        context.user_data["current_mahsulot"] = text
        context.user_data["waiting_for_mahsulot"] = False
        context.user_data["waiting_for_rasm"] = True
        await update.message.reply_text(
            "Тош учун расм юборинг ёки /rasmsiz буйруғини юборинг:",
            reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
        )
        return

    if context.user_data.get("waiting_for_qalinlik"):
        qalinlik = text.replace(",", ".")
        if not re.match(r"^\d+(\.\d+)?$", qalinlik):
            await update.message.reply_text(
                "Илтимос, тўғри қалинлик киритинг (масалан, 1.3 ёки 1,3)!",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
            return
        if qalinlik not in data["qalinlik"]:
            data["qalinlik"].append(qalinlik)
            save_data(data)
        context.user_data["waiting_for_qalinlik"] = False
        await update.message.reply_text(
            f"Қалинлик '{qalinlik}' сақланди!",
            reply_markup=ReplyKeyboardMarkup(ADMIN_KEYBOARD, resize_keyboard=True)
        )
        return

    if context.user_data.get("waiting_for_xodim_id"):
        if not text.isdigit():
            await update.message.reply_text(
                "Илтимос, фақат рақамли ID киритинг!",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
            return
        context.user_data["current_xodim_id"] = text
        context.user_data["waiting_for_xodim_id"] = False
        context.user_data["waiting_for_xodim_nom"] = True
        await update.message.reply_text(
            "Ходимнинг номини киритинг:",
            reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
        )
        return

    if context.user_data.get("waiting_for_xodim_nom"):
        xodim_id = context.user_data["current_xodim_id"]
        user_info = await context.bot.get_chat(xodim_id)
        data["xodimlar"][xodim_id] = {
            "nom": text,
            "username": user_info.username or "",
            "profil_nomi": user_info.first_name or "",
            "qoshilgan_vaqt": datetime.now(tz=TASHKENT_TZ).strftime("%H:%M %d.%m.%Y")
        }
        save_data(data)
        context.user_data["waiting_for_xodim_nom"] = False
        context.user_data["current_xodim_id"] = None
        await update.message.reply_text(
            f"Ходим '{text}' сақланди!",
            reply_markup=ReplyKeyboardMarkup(ADMIN_KEYBOARD, resize_keyboard=True)
        )
        return

    if text == "Ҳисобот":
        context.user_data["waiting_for_report"] = True
        context.user_data["previous_menu"] = None
        await update.message.reply_text(
            "Ҳисобот турини танланг:",
            reply_markup=ReplyKeyboardMarkup(REPORT_KEYBOARD, resize_keyboard=True)
        )
        return

    if context.user_data.get("waiting_for_report"):
        if text == "Бугунги":
            await generate_report(update, context, data)
            context.user_data["waiting_for_report"] = False
            await update.message.reply_text(
                "Админ панели",
                reply_markup=ReplyKeyboardMarkup(ADMIN_KEYBOARD, resize_keyboard=True)
            )
        elif text == "Муддат оралиғи":
            context.user_data["waiting_for_report"] = False
            await handle_date_range(update, context)
        return

    if text == "Ўчириш":
        context.user_data["waiting_for_delete"] = True
        context.user_data["previous_menu"] = None
        await update.message.reply_text(
            "Нималарни ўчирмоқчисиз?",
            reply_markup=ReplyKeyboardMarkup(DELETE_KEYBOARD, resize_keyboard=True)
        )
        return

    if context.user_data.get("waiting_for_delete"):
        if text == "🗑 Обработка":
            if not data["obrabotka"]:
                await update.message.reply_text("Обработкалар мавжуд эмас!")
                return
            keyboard = [[InlineKeyboardButton(name, callback_data=f"delete_obrabotka_{name}")] for name in sorted(data["obrabotka"])]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await update.message.reply_text("Ўчириладиган обработкани танланг:", reply_markup=reply_markup)
        elif text == "🗑 Каталог":
            if not data["katalog"]:
                await update.message.reply_text("Каталоглар мавжуд эмас!")
                return
            keyboard = [[InlineKeyboardButton(name, callback_data=f"delete_katalog_{name}")] for name in sorted(data["katalog"])]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await update.message.reply_text("Ўчириладиган каталогни танланг:", reply_markup=reply_markup)
        elif text == "🗑 Станок":
            if not data["obrabotka"]:
                await update.message.reply_text("Обработкалар мавжуд эмас!")
                return
            keyboard = [[InlineKeyboardButton(obrabotka, callback_data=f"delete_stanok_obrabotka_{obrabotka}")] for obrabotka in sorted(data["obrabotka"])]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await update.message.reply_text("Станок ўчириладиган обработкани танланг:", reply_markup=reply_markup)
        elif text == "🗑 Қалинлик":
            if not data["qalinlik"]:
                await update.message.reply_text("Қалинликлар мавжуд эмас!")
                return
            keyboard = [[InlineKeyboardButton(q, callback_data=f"delete_qalinlik_{q}")] for q in sorted(data["qalinlik"], key=lambda x: float(x.replace(",", ".")))]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await update.message.reply_text("Ўчириладиган қалинликни танланг:", reply_markup=reply_markup)
        elif text == "🗑 Тош":
            if not data["katalog"]:
                await update.message.reply_text("Каталоглар мавжуд эмас!")
                return
            keyboard = [[InlineKeyboardButton(name, callback_data=f"delete_mahsulot_katalog_{name}")] for name in sorted(data["katalog"])]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await update.message.reply_text("Тош ўчириладиган каталогни танланг:", reply_markup=reply_markup)
        return

    if text == "+ Обработка":
        if data["obrabotka"]:
            keyboard = [[InlineKeyboardButton(name, callback_data=f"obrabotka_{name}")] for name in sorted(data["obrabotka"])]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await update.message.reply_text("Мавжуд обработкалар:", reply_markup=reply_markup)
        context.user_data["waiting_for_obrabotka"] = True
        await update.message.reply_text(
            "Янги обработка номини киритинг:",
            reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
        )
    elif text == "+ Каталог":
        context.user_data["previous_menu"] = "katalog"
        await show_katalog(update, context, admin=True)
    elif text == "Каталог қўшиш":
        context.user_data["waiting_for_katalog"] = True
        await update.message.reply_text(
            "Янги каталог номини киритинг:",
            reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
        )
    elif text == "+ Қалинлик":
        context.user_data["waiting_for_qalinlik"] = True
        await update.message.reply_text(
            "Қалинликни киритинг (масалан, 1.3 ёки 1,3):",
            reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
        )
    elif text == "+ Ходим":
        if data["xodimlar"]:
            for xodim_id, info in data["xodimlar"].items():
                username = info.get("username", "Йўқ")
                if username and not username.startswith("@"):
                    username = f"@{username}"
                await context.bot.send_message(
                    chat_id=update.message.chat_id,
                    text=(
                        f"Ходим: {info['nom']}\n"
                        f"—————————\n"
                        f"Профиль номи: {info['profil_nomi'] or 'Йўқ'}\n"
                        f"Усернаме: {username}\n"
                        f"ID рақами: {xodim_id}\n"
                        f"—————————\n"
                        f"{info['qoshilgan_vaqt']} дан бери"
                    ),
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Ходимни ўчириш", callback_data=f"delete_xodim_{xodim_id}")]])
                )
        context.user_data["waiting_for_xodim_id"] = True
        await update.message.reply_text(
            "Ходим ID рақамини киритинг:",
            reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
        )
    else:
        await update.message.reply_text(
            "Тугмалардан бирини танланг.",
            reply_markup=ReplyKeyboardMarkup(ADMIN_KEYBOARD, resize_keyboard=True)
        )

async def handle_xodim_message(update: Update, context: ContextTypes.DEFAULT_TYPE, data, text, user_id):
    if text == "Орқага қайтиш":
        if context.user_data.get("waiting_for_smena"):
            context.user_data.clear()
            await update.message.reply_text(
                "Ходим панелига қайтиш:",
                reply_markup=ReplyKeyboardMarkup(XODIM_KEYBOARD, resize_keyboard=True)
            )
        elif context.user_data.get("waiting_for_obrabotka"):
            context.user_data["waiting_for_obrabotka"] = False
            keyboard = [
                [InlineKeyboardButton("Кун", callback_data="smena_Кун")],
                [InlineKeyboardButton("Тун", callback_data="smena_Тун")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            context.user_data["last_message_id"] = (await update.message.reply_text(
                "Сменани танланг:",
                reply_markup=reply_markup
            )).message_id
        elif context.user_data.get("waiting_for_katalog"):
            context.user_data["waiting_for_katalog"] = False
            keyboard = [[InlineKeyboardButton(name, callback_data=f"order_obrabotka_{name}")] for name in sorted(data["obrabotka"])]
            reply_markup = InlineKeyboardMarkup(keyboard)
            context.user_data["last_message_id"] = (await update.message.reply_text(
                "Обработкалардан бирини танланг:",
                reply_markup=reply_markup
            )).message_id
        elif context.user_data.get("waiting_for_qalinlik"):
            context.user_data["waiting_for_qalinlik"] = False
            keyboard = [[InlineKeyboardButton(mahsulot["nom"], callback_data=f"order_mahsulot_{mahsulot['id']}")] for mahsulot in sorted(data["katalog"][context.user_data["order"]["katalog"]]["items"], key=lambda x: x["nom"])]
            reply_markup = InlineKeyboardMarkup(keyboard)
            context.user_data["last_message_id"] = (await update.message.reply_text(
                "Тошлардан бирини танланг:",
                reply_markup=reply_markup
            )).message_id
        elif context.user_data.get("waiting_for_manual_qalinlik"):
            context.user_data["waiting_for_manual_qalinlik"] = False
            keyboard = [[InlineKeyboardButton(mahsulot["nom"], callback_data=f"order_mahsulot_{mahsulot['id']}")] for mahsulot in sorted(data["katalog"][context.user_data["order"]["katalog"]]["items"], key=lambda x: x["nom"])]
            reply_markup = InlineKeyboardMarkup(keyboard)
            context.user_data["last_message_id"] = (await update.message.reply_text(
                "Тошлардан бирини танланг:",
                reply_markup=reply_markup
            )).message_id
        elif context.user_data.get("waiting_for_eni"):
            context.user_data["waiting_for_eni"] = False
            keyboard = [
                [InlineKeyboardButton(f"{q} см", callback_data=f"qalinlik_{q}")] for q in sorted(data["qalinlik"], key=lambda x: float(x.replace(",", ".")))
            ] + [[InlineKeyboardButton("Қўлда киритиш", callback_data="manual_qalinlik")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            context.user_data["last_message_id"] = (await update.message.reply_text(
                "Қалинликни танланг:",
                reply_markup=reply_markup
            )).message_id
        elif context.user_data.get("waiting_for_boy"):
            context.user_data["waiting_for_boy"] = False
            keyboard = [
                [InlineKeyboardButton("размер", callback_data="размер")],
                [InlineKeyboardButton("произвол", callback_data="произвол")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            context.user_data["last_message_id"] = (await update.message.reply_text(
                "Кераклисини танланг:",
                reply_markup=reply_markup
            )).message_id
        elif context.user_data.get("waiting_for_son") or context.user_data.get("waiting_for_metr"):
            context.user_data["waiting_for_son"] = False
            context.user_data["waiting_for_metr"] = False
            await update.message.reply_text(
                "Энини киритинг (сантиметрда):",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
        elif context.user_data.get("waiting_for_izoh"):
            context.user_data["waiting_for_izoh"] = False
            if context.user_data["order"].get("type") == "размер":
                context.user_data["waiting_for_son"] = True
                await update.message.reply_text(
                    "Сонини киритинг:",
                    reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
                )
            else:
                context.user_data["waiting_for_metr"] = True
                await update.message.reply_text(
                    "Метрни киритинг (метрда):",
                    reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
                )
        else:
            context.user_data.clear()
            await update.message.reply_text(
                "Ходим панелига қайтиш:",
                reply_markup=ReplyKeyboardMarkup(XODIM_KEYBOARD, resize_keyboard=True)
            )
        return

    if context.user_data.get("waiting_for_manual_qalinlik"):
        qalinlik = text.replace(",", ".")
        if not re.match(r"^\d+(\.\d+)?$", qalinlik):
            await update.message.reply_text(
                "Илтимос, тўғри қалинлик киритинг (масалан, 1.3 ёки 1,3)!",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
            return
        context.user_data["order"]["qalinlik"] = qalinlik
        context.user_data["waiting_for_manual_qalinlik"] = False
        keyboard = [
            [InlineKeyboardButton("размер", callback_data="размер")],
            [InlineKeyboardButton("произвол", callback_data="произвол")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.user_data["last_message_id"] = (await update.message.reply_text(
            "Кераклисини танланг:",
            reply_markup=reply_markup
        )).message_id
        return

    if context.user_data.get("waiting_for_eni"):
        if not text.isdigit():
            await update.message.reply_text(
                "Илтимос, фақат рақам киритинг (сантиметрда)!",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
            return
        context.user_data["order"]["eni"] = text
        if context.user_data["order"]["type"] == "размер":
            context.user_data["waiting_for_eni"] = False
            context.user_data["waiting_for_boy"] = True
            await update.message.reply_text(
                "Бўйини киритинг (сантиметрда):",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
        else:
            context.user_data["waiting_for_eni"] = False
            context.user_data["waiting_for_metr"] = True
            await update.message.reply_text(
                "Метрни киритинг (метрда):",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
        return

    if context.user_data.get("waiting_for_boy"):
        if not text.isdigit():
            await update.message.reply_text(
                "Илтимос, фақат рақам киритинг (сантиметрда)!",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
            return
        context.user_data["order"]["boy"] = text
        context.user_data["waiting_for_boy"] = False
        context.user_data["waiting_for_son"] = True
        await update.message.reply_text(
            "Сонини киритинг:",
            reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
        )
        return

    if context.user_data.get("waiting_for_son"):
        if not text.isdigit():
            await update.message.reply_text(
                "Илтимос, фақат рақам киритинг!",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
            return
        context.user_data["order"]["son"] = text
        context.user_data["waiting_for_son"] = False
        context.user_data["waiting_for_izoh"] = True
        await update.message.reply_text(
            "Изоҳ киритасизми?",
            reply_markup=ReplyKeyboardMarkup(ORDER_KEYBOARD, resize_keyboard=True)
        )
        return

    if context.user_data.get("waiting_for_metr"):
        try:
            metr = float(text.replace(",", "."))
            if metr <= 0:
                raise ValueError
        except ValueError:
            await update.message.reply_text(
                "Илтимос, тўғри метр қийматини киритинг (масалан, 1.5)!",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
            return
        context.user_data["order"]["metr"] = str(metr)
        context.user_data["waiting_for_metr"] = False
        context.user_data["waiting_for_izoh"] = True
        await update.message.reply_text(
            "Изоҳ киритасизми?",
            reply_markup=ReplyKeyboardMarkup(ORDER_KEYBOARD, resize_keyboard=True)
        )
        return

    if context.user_data.get("waiting_for_izoh"):
        context.user_data["order"]["izoh"] = text
        context.user_data["order"]["xodim_id"] = str(user_id)
        context.user_data["order"]["vaqt"] = datetime.now(tz=TASHKENT_TZ).strftime("%d.%m.%Y / %H:%M")
        data["orders"].append(context.user_data["order"])
        save_data(data)
        await append_to_excel(context.user_data["order"], data)
        context.user_data.clear()
        await update.message.reply_text(
            "Муваффақиятли қўшилди!",
            reply_markup=ReplyKeyboardMarkup(XODIM_KEYBOARD, resize_keyboard=True)
        )
        return

    if text == "Янги қўшиш":
        context.user_data["order"] = {}
        context.user_data["waiting_for_smena"] = True
        keyboard = [
            [InlineKeyboardButton("Кун", callback_data="smena_Кун")],
            [InlineKeyboardButton("Тун", callback_data="smena_Тун")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.user_data["last_message_id"] = (await update.message.reply_text(
            "Сменани танланг:",
            reply_markup=reply_markup
        )).message_id
    elif text == "Изоҳсиз":
        if "order" not in context.user_data or not context.user_data.get("order"):
            await update.message.reply_text(
                "Илтимос, аввал буюртма маълумотларини тўлиқ киритинг! 'Янги қўшиш' тугмасидан бошланг.",
                reply_markup=ReplyKeyboardMarkup(XODIM_KEYBOARD, resize_keyboard=True)
            )
            return
        context.user_data["order"]["izoh"] = "Изоҳсиз"
        context.user_data["order"]["xodim_id"] = str(user_id)
        context.user_data["order"]["vaqt"] = datetime.now(tz=TASHKENT_TZ).strftime("%d.%m.%Y / %H:%M")
        data["orders"].append(context.user_data["order"])
        save_data(data)
        await append_to_excel(context.user_data["order"], data)
        context.user_data.clear()
        await update.message.reply_text(
            "Муваффақиятли қўшилди!",
            reply_markup=ReplyKeyboardMarkup(XODIM_KEYBOARD, resize_keyboard=True)
        )
    else:
        await update.message.reply_text(
            "Тугмалардан бирини танланг.",
            reply_markup=ReplyKeyboardMarkup(XODIM_KEYBOARD, resize_keyboard=True)
        )

async def show_katalog(update: Update, context: ContextTypes.DEFAULT_TYPE, admin=False):
    data = load_data()
    if not data["katalog"]:
        await update.message.reply_text(
            "Ҳозирча каталоглар йўқ.",
            reply_markup=ReplyKeyboardMarkup(KATALOG_KEYBOARD, resize_keyboard=True)
        )
        return
    keyboard = [[InlineKeyboardButton(name, callback_data=f"katalog_{name}")] for name in sorted(data["katalog"])]
    reply_markup = InlineKeyboardMarkup(keyboard)
    context.user_data["last_message_id"] = (await update.message.reply_text(
        "Каталоглардан бирини танланг:",
        reply_markup=reply_markup
    )).message_id
    if admin:
        await update.message.reply_text(
            "Ёки янги каталог қўшинг:",
            reply_markup=ReplyKeyboardMarkup(KATALOG_KEYBOARD, resize_keyboard=True)
        )

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    data = load_data()
    
    if user_id != ADMIN_ID:
        await update.message.reply_text("Кечирасиз, бу фақат админ учун!")
        return
    
    if context.user_data.get("waiting_for_rasm"):
        photo = update.message.photo[-1]
        katalog = context.user_data["current_katalog"]
        mahsulot = context.user_data["current_mahsulot"]
        mahsulot_id = str(uuid.uuid4())
        
        photo_filename = f"mahsulot_{mahsulot_id}.jpg"
        photo_path = os.path.join(os.path.dirname(__file__), photo_filename)
        
        file = await context.bot.get_file(photo.file_id)
        async with aiohttp.ClientSession() as session:
            async with session.get(file.file_path) as resp:
                if resp.status == 200:
                    async with aiofiles.open(photo_path, "wb") as f:
                        await f.write(await resp.read())
    
        if katalog not in data["katalog"]:
            data["katalog"][katalog] = {"id": str(uuid.uuid4()), "items": []}
        data["katalog"][katalog]["items"].append({"nom": mahsulot, "rasm_path": photo_filename, "id": mahsulot_id})
        save_data(data)
    
        context.user_data["waiting_for_rasm"] = False
        context.user_data["current_katalog"] = None
        context.user_data["current_mahsulot"] = None
        await update.message.reply_text(
            f"Тош '{mahsulot}' расм билан сақланди!",
            reply_markup=ReplyKeyboardMarkup(KATALOG_KEYBOARD, resize_keyboard=True)
        )

async def append_to_sheets(order, data):
    try:
        service = get_sheets_service()
        if not service:
            print("Google Sheets xizmatiga ulanib bo‘lmadi.")
            return

        sana_vaqt = datetime.now(tz=TASHKENT_TZ).strftime("%d.%m.%Y / %H:%M")
        smena = order.get("smena", "")
        stanok = order.get("stanok", "")
        mahsulot = order.get("mahsulot", "")
        qalinlik = f"{order.get('qalinlik', '').replace('.', ',')} см" if order.get("qalinlik") else ""
        obrabotka = order.get("obrabotka", "")
        
        if order.get("type") == "размер":
            razmer = f"{order.get('eni', '')}x{order.get('boy', '')} см"
            miqdor = f"{order.get('son', '')} дона"
            eni = float(order.get("eni", 0))
            boy = float(order.get("boy", 0))
            son = int(order.get("son", 0))
            metr_kv = (eni * boy * son) / 10000
        else:
            razmer = f"{order.get('eni', '')} см произвол"
            miqdor = f"{order.get('metr', '')} metr"
            eni = float(order.get("eni", 0))
            metr = float(order.get("metr", 0))
            metr_kv = (eni * metr) / 100

        izoh = order.get("izoh", "")
        hisobga_olinmaydi = ""
        tun_kunlik = metr_kv if smena == "Тун" else 0
        kun_kunlik = metr_kv if smena == "Кун" else 0
        xodim_nom = data["xodimlar"].get(order.get("xodim_id", ""), {}).get("nom", "")

        values = [
            [
                sana_vaqt, smena, stanok, mahsulot, qalinlik,
                obrabotka, razmer, miqdor, metr_kv, izoh,
                hisobga_olinmaydi, tun_kunlik, kun_kunlik, xodim_nom
            ]
        ]

        body = {"values": values}
        result = service.spreadsheets().values().append(
            spreadsheetId=SPREADSHEET_ID,
            range=f"{SHEET_NAME}!A:A",
            valueInputOption="RAW",
            insertDataOption="INSERT_ROWS",
            body=body
        ).execute()

        print(f"Google Sheets’ga qo‘shildi: {result}")

        sheet = service.spreadsheets().values().get(
            spreadsheetId=SPREADSHEET_ID,
            range=f"{SHEET_NAME}!A:N"
        ).execute()
        rows = sheet.get("values", [])
        last_row = len(rows) if rows else 1

        formulas = [
            [f"=SUM(I2:I{last_row})"],
            [f"=SUM(L2:L{last_row})"],
            [f"=SUM(M2:M{last_row})"]
        ]
        body = {"values": formulas}
        service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID,
            range=f"{SHEET_NAME}!O2:O4",
            valueInputOption="USER_ENTERED",
            body=body
        ).execute()

        print(f"Summalar {SHEET_NAME}!O2:O4 ga yozildi")

    except Exception as e:
        print(f"Google Sheets’ga yozishda xato: {e}")

async def append_to_excel(order, data):
    init_excel()
    wb = openpyxl.load_workbook(REPORT_FILE)
    ws = wb.active
    
    ws.insert_rows(4)
    row = 4
    
    sana_vaqt = datetime.now(tz=TASHKENT_TZ).strftime("%d.%m.%Y / %H:%M")
    ws[f'A{row}'] = sana_vaqt
    
    ws[f'B{row}'] = order.get("smena", "")
    ws[f'C{row}'] = order.get("stanok", "")
    ws[f'D{row}'] = order.get("mahsulot", "")
    ws[f'E{row}'] = f"{order.get('qalinlik', '').replace('.', ',')} см" if order.get("qalinlik") else ""
    ws[f'F{row}'] = order.get("obrabotka", "")
    
    if order.get("type") == "размер":
        ws[f'G{row}'] = f"{order.get('eni', '')}x{order.get('boy', '')} см"
        ws[f'H{row}'] = f"{order.get('son', '')} дона"
        eni = float(order.get("eni", 0))
        boy = float(order.get("boy", 0))
        son = int(order.get("son", 0))
        metr_kv = (eni * boy * son) / 10000
    else:
        ws[f'G{row}'] = f"{order.get('eni', '')} см произвол"
        ws[f'H{row}'] = f"{order.get('metr', '')} metr"
        eni = float(order.get("eni", 0))
        metr = float(order.get("metr", 0))
        metr_kv = (eni * metr) / 100
    
    ws[f'I{row}'] = metr_kv
    ws[f'J{row}'] = order.get("izoh", "")
    
    if order.get("smena") == "Кун":
        ws[f'M{row}'] = metr_kv
        ws[f'L{row}'] = 0
    else:
        ws[f'L{row}'] = metr_kv
        ws[f'M{row}'] = 0
    
    xodim_id = order.get("xodim_id", "")
    ws[f'N{row}'] = data["xodimlar"].get(xodim_id, {}).get("nom", "")
    
    black_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    for col in range(1, 15):
        cell = ws.cell(row=row, column=col)
        cell.border = black_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    last_row = ws.max_row
    ws[f'H{last_row+1}'] = "Jami:"
    ws[f'I{last_row+1}'] = f'=SUM(I4:I{last_row})'
    ws[f'L{last_row+1}'] = f'=SUM(L4:L{last_row})'
    ws[f'M{last_row+1}'] = f'=SUM(M4:M{last_row})'
    
    for col in range(1, 15):
        cell = ws.cell(row=last_row+1, column=col)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    wb.save(REPORT_FILE)
    
    await append_to_sheets(order, data)

async def generate_report(update: Update, context: ContextTypes.DEFAULT_TYPE, data, start_date=None, end_date=None):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ҳисобот"
        
        headers = [
            "Сана", "Смена", "Станок", "Тош номи", "Қалинлик",
            "Обработка", "Размери", "Миқдори", "Хажми(м²)", "Изоҳ",
            "Ҳисобга олинмайди", "Тун кунлик", "Кун кунлик", "Ходим номи"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        gray_fill = PatternFill(start_color="31869B", end_color="31869B", fill_type="solid")
        black_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=3, column=col)
            cell.fill = gray_fill
            cell.border = black_border
        
        column_widths = [130, 76, 95, 250, 60, 80, 120, 120, 95, 80, 80, 80, 80, 80]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width / 7
        
        row = 4
        today = datetime.now(tz=TASHKENT_TZ).date()
        orders = sorted(data["orders"], key=lambda x: datetime.strptime(x["vaqt"], "%d.%m.%Y / %H:%M"), reverse=True)
        
        for order in orders:
            try:
                order_date = datetime.strptime(order["vaqt"], "%d.%m.%Y / %H:%M").date()
                
                if start_date is None and end_date is None:
                    if order_date != today:
                        continue
                elif start_date and end_date:
                    if not (start_date <= order_date <= end_date):
                        continue
                
                ws[f'A{row}'] = order.get("vaqt", "")
                ws[f'B{row}'] = order.get("smena", "")
                ws[f'C{row}'] = order.get("stanok", "")
                ws[f'D{row}'] = order.get("mahsulot", "")
                ws[f'E{row}'] = f"{order.get('qalinlik', '').replace('.', ',')} см" if order.get("qalinlik") else ""
                ws[f'F{row}'] = order.get("obrabotka", "")
                
                if order.get("type") == "размер":
                    ws[f'G{row}'] = f"{order.get('eni', '')}x{order.get('boy', '')} см"
                    ws[f'H{row}'] = f"{order.get('son', '')} дона"
                    eni = float(order.get("eni", 0))
                    boy = float(order.get("boy", 0))
                    son = int(order.get("son", 0))
                    metr_kv = (eni * boy * son) / 10000
                else:
                    ws[f'G{row}'] = f"{order.get('eni', '')} см произвол"
                    ws[f'H{row}'] = f"{order.get('metr', '')} metr"
                    eni = float(order.get("eni", 0))
                    metr = float(order.get("metr", 0))
                    metr_kv = (eni * metr) / 100
                
                ws[f'I{row}'] = metr_kv
                ws[f'J{row}'] = order.get("izoh", "")
                
                if order.get("smena") == "Кун":
                    ws[f'M{row}'] = metr_kv
                    ws[f'L{row}'] = 0
                else:
                    ws[f'L{row}'] = metr_kv
                    ws[f'M{row}'] = 0
                
                xodim_id = order.get("xodim_id", "")
                ws[f'N{row}'] = data["xodimlar"].get(xodim_id, {}).get("nom", "")
                
                for col in range(1, 15):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = black_border
                
                row += 1
            except Exception as e:
                print(f"Қаторни қайта ишлашда хато: {e}")
                continue
        
        if row > 4:
            ws[f'H{row}'] = "Жами:"
            ws[f'I{row}'] = f'=SUM(I4:I{row-1})'
            ws[f'L{row}'] = f'=SUM(L4:L{row-1})'
            ws[f'M{row}'] = f'=SUM(M4:M{row-1})'
            
            for col in range(1, 15):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = black_border
        
        temp_file = "temp_report.xlsx"
        wb.save(temp_file)
        
        if start_date and end_date:
            filename = f"Ҳисобот_{start_date.strftime('%d_%m_%Y')}_{end_date.strftime('%d_%m_%Y')}.xlsx"
        else:
            filename = f"Ҳисобот_{datetime.now(tz=TASHKENT_TZ).strftime('%d_%m_%Y')}.xlsx"
        
        if update:
            await update.message.reply_document(
                document=open(temp_file, "rb"),
                filename=filename
            )
        else:
            await context.bot.send_document(
                chat_id=ADMIN_ID,
                document=open(temp_file, "rb"),
                filename=filename
            )
        os.remove(temp_file)
    
    except Exception as e:
        print(f"Ҳисобот генерациясида хато: {e}")
        error_message = f"Ҳисоботни чиқаришда хато юз берди: {str(e)}"
        if update:
            await update.message.reply_text(error_message)
        else:
            await context.bot.send_message(chat_id=ADMIN_ID, text=error_message)

async def handle_date_range(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Бошланғич санани киритинг (кун.ой.йил форматида, масалан: 01.01.2023):",
        reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
    )
    context.user_data["waiting_for_start_date"] = True

async def process_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    data = load_data()
    
    try:
        if context.user_data.get("waiting_for_start_date"):
            start_date = datetime.strptime(text, "%d.%m.%Y").date()
            context.user_data["start_date"] = start_date
            context.user_data["waiting_for_start_date"] = False
            context.user_data["waiting_for_end_date"] = True
            await update.message.reply_text(
                "Якуний санани киритинг (кун.ой.йил форматида):",
                reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
            )
        elif context.user_data.get("waiting_for_end_date"):
            end_date = datetime.strptime(text, "%d.%m.%Y").date()
            context.user_data["end_date"] = end_date
            context.user_data["waiting_for_end_date"] = False
            await generate_report(
                update, 
                context, 
                data, 
                context.user_data["start_date"], 
                context.user_data["end_date"]
            )
            await update.message.reply_text(
                "Админ панели",
                reply_markup=ReplyKeyboardMarkup(ADMIN_KEYBOARD, resize_keyboard=True)
            )
    except ValueError:
        await update.message.reply_text(
            "Нотўғри сана формати! Илтимос, кун.ой.йил форматида киритинг (масалан: 01.01.2023)",
            reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
        )

async def send_daily_report(context: ContextTypes.DEFAULT_TYPE):
    data = load_data()
    today = datetime.now(tz=TASHKENT_TZ).date()
    await generate_report(None, context, data, today, today)

async def send_10day_report(context: ContextTypes.DEFAULT_TYPE):
    today = datetime.now(tz=TASHKENT_TZ).date()
    if today.day in [10, 20, 30]:
        data = load_data()
        start_date = today - timedelta(days=9)
        await generate_report(None, context, data, start_date, today)
        data["orders"] = [order for order in data["orders"] 
                         if datetime.strptime(order["vaqt"], "%d.%m.%Y / %H:%M").date() >= today]
        save_data(data)

async def button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = load_data()

    if context.user_data.get("last_message_id"):
        try:
            await query.message.delete()
        except:
            pass

    if query.data.startswith("smena_"):
        smena = query.data.replace("smena_", "").capitalize()
        context.user_data["order"]["smena"] = smena
        context.user_data["waiting_for_smena"] = False
        context.user_data["waiting_for_obrabotka"] = True
        keyboard = [[InlineKeyboardButton(name, callback_data=f"order_obrabotka_{name}")] for name in sorted(data["obrabotka"])]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.user_data["last_message_id"] = (await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Обработкалардан бирини танланг:",
            reply_markup=reply_markup
        )).message_id
        return

    if query.data.startswith("obrabotka_"):
        obrabotka = query.data.replace("obrabotka_", "")
        context.user_data["current_obrabotka"] = obrabotka
        context.user_data["waiting_for_stanok"] = True
        keyboard = [[InlineKeyboardButton(stanok, callback_data=f"stanok_{stanok}")] for stanok in sorted(data["obrabotka"][obrabotka])]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.user_data["last_message_id"] = (await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=f"Станоклар ({obrabotka}):",
            reply_markup=reply_markup
        )).message_id
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Янги станок номини киритинг:",
            reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
        )
        return

    if query.data.startswith("katalog_"):
        katalog = query.data.replace("katalog_", "")
        if katalog not in data["katalog"]:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Каталог топилмади!"
            )
            return
        context.user_data["current_katalog"] = katalog
        context.user_data["waiting_for_mahsulot"] = True
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Тош номини киритинг:",
            reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
        )
        return

    if query.data.startswith("delete_xodim_"):
        xodim_id = query.data.replace("delete_xodim_", "")
        if xodim_id in data["xodimlar"]:
            del data["xodimlar"][xodim_id]
            save_data(data)
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Ходим ўчирилди!"
            )
        else:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Ходим топилмади!"
            )
        return

    if query.data.startswith("rasm_katalog_"):
        katalog_name = query.data.replace("rasm_katalog_", "")
        if katalog_name not in data["katalog"] or not data["katalog"][katalog_name]["items"]:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text=f"'{katalog_name}' каталог ичида маҳсулотлар йўқ!"
            )
            return
        keyboard = [[InlineKeyboardButton(mahsulot["nom"], callback_data=f"rasm_mahsulot_{katalog_name}_{mahsulot['id']}")] for mahsulot in sorted(data["katalog"][katalog_name]["items"], key=lambda x: x["nom"])]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.user_data["last_message_id"] = (await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=f"'{katalog_name}' каталог ичидаги тошлар:",
            reply_markup=reply_markup
        )).message_id
        return

    if query.data.startswith("rasm_mahsulot_"):
        _, katalog_name, mahsulot_id = query.data.split("_", 2)
        if katalog_name not in data["katalog"]:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Каталог топилмади!"
            )
            return
        mahsulot = next((item for item in data["katalog"][katalog_name]["items"] if item["id"] == mahsulot_id), None)
        if not mahsulot:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Тош топилмади!"
            )
            return
        if mahsulot.get("rasm_path"):
            try:
                await context.bot.send_photo(
                    chat_id=query.message.chat_id,
                    photo=open(mahsulot["rasm_path"], "rb"),
                    caption=mahsulot["nom"]
                )
            except FileNotFoundError:
                await context.bot.send_message(
                    chat_id=query.message.chat_id,
                    text=f"'{mahsulot['nom']}' тоши учун расм топилмади!"
                )
        else:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text=f"'{mahsulot['nom']}' тоши учун расм мавжуд эмас."
            )
        keyboard = [[InlineKeyboardButton(name, callback_data=f"rasm_katalog_{name}")] for name in sorted(data["katalog"])]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.user_data["last_message_id"] = (await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Яна каталог танланг:",
            reply_markup=reply_markup
        )).message_id
        return

    if query.data.startswith("order_obrabotka_"):
        obrabotka = query.data.replace("order_obrabotka_", "")
        if obrabotka not in data["obrabotka"]:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Обработка топилмади!"
            )
            return
        context.user_data["order"]["obrabotka"] = obrabotka
        context.user_data["waiting_for_obrabotka"] = False
        keyboard = [[InlineKeyboardButton(stanok, callback_data=f"order_stanok_{stanok}")] for stanok in sorted(data["obrabotka"][obrabotka])]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.user_data["last_message_id"] = (await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Станоклардан бирини танланг:",
            reply_markup=reply_markup
        )).message_id
        return

    if query.data.startswith("order_stanok_"):
        stanok = query.data.replace("order_stanok_", "")
        obrabotka = context.user_data["order"].get("obrabotka")
        if obrabotka not in data["obrabotka"] or stanok not in data["obrabotka"][obrabotka]:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Станок топилмади!"
            )
            return
        context.user_data["order"]["stanok"] = stanok
        context.user_data["waiting_for_katalog"] = True
        keyboard = [[InlineKeyboardButton(name, callback_data=f"order_katalog_{name}")] for name in sorted(data["katalog"])]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.user_data["last_message_id"] = (await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Каталоглардан бирини танланг:",
            reply_markup=reply_markup
        )).message_id
        return

    if query.data.startswith("order_katalog_"):
        katalog_name = query.data.replace("order_katalog_", "")
        if katalog_name not in data["katalog"]:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Каталог топилмади!"
            )
            return
        context.user_data["order"]["katalog"] = katalog_name
        context.user_data["waiting_for_katalog"] = False
        keyboard = [[InlineKeyboardButton(mahsulot["nom"], callback_data=f"order_mahsulot_{mahsulot['id']}")] for mahsulot in sorted(data["katalog"][katalog_name]["items"], key=lambda x: x["nom"])]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.user_data["last_message_id"] = (await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Тошлардан бирини танланг:",
            reply_markup=reply_markup
        )).message_id
        return

    if query.data.startswith("order_mahsulot_"):
        mahsulot_id = query.data.replace("order_mahsulot_", "")
        katalog_name = context.user_data["order"].get("katalog")
        if katalog_name not in data["katalog"]:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Каталог топилмади!"
            )
            return
        mahsulot = next((item for item in data["katalog"][katalog_name]["items"] if item["id"] == mahsulot_id), None)
        if not mahsulot:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Тош топилмади!"
            )
            return
        context.user_data["order"]["mahsulot"] = mahsulot["nom"]
        context.user_data["waiting_for_qalinlik"] = True
        keyboard = [
            [InlineKeyboardButton(f"{q} см", callback_data=f"qalinlik_{q}")] for q in sorted(data["qalinlik"], key=lambda x: float(x.replace(",", ".")))
        ] + [[InlineKeyboardButton("Қўлда киритиш", callback_data="manual_qalinlik")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.user_data["last_message_id"] = (await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Қалинликни танланг:",
            reply_markup=reply_markup
        )).message_id
        return

    if query.data.startswith("qalinlik_"):
        qalinlik = query.data.replace("qalinlik_", "")
        context.user_data["order"]["qalinlik"] = qalinlik
        context.user_data["waiting_for_qalinlik"] = False
        keyboard = [
            [InlineKeyboardButton("размер", callback_data="размер")],
            [InlineKeyboardButton("произвол", callback_data="произвол")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.user_data["last_message_id"] = (await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Кераклисини танланг:",
            reply_markup=reply_markup
        )).message_id
        return

    if query.data == "manual_qalinlik":
        context.user_data["waiting_for_qalinlik"] = False
        context.user_data["waiting_for_manual_qalinlik"] = True
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Қалинликни киритинг (масалан, 1.3 ёки 1,3):",
            reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
        )
        return

    if query.data in ["размер", "произвол"]:
        context.user_data["order"]["type"] = query.data
        context.user_data["waiting_for_eni"] = True
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Энини киритинг (сантиметрда):",
            reply_markup=ReplyKeyboardMarkup([["Орқага қайтиш"]], resize_keyboard=True)
        )
        return

    if query.data.startswith("delete_obrabotka_"):
        obrabotka = query.data.replace("delete_obrabotka_", "")
        if obrabotka in data["obrabotka"]:
            del data["obrabotka"][obrabotka]
            save_data(data)
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text=f"Обработка '{obrabotka}' ўчирилди!"
            )
        else:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Обработка топилмади!"
            )
        return

    if query.data.startswith("delete_katalog_"):
        katalog = query.data.replace("delete_katalog_", "")
        if katalog in data["katalog"]:
            for item in data["katalog"][katalog]["items"]:
                if item.get("rasm_path") and os.path.exists(item["rasm_path"]):
                    os.remove(item["rasm_path"])
            del data["katalog"][katalog]
            save_data(data)
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text=f"Каталог '{katalog}' ўчирилди!"
            )
        else:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Каталог топилмади!"
            )
        return

    if query.data.startswith("delete_stanok_obrabotka_"):
        obrabotka = query.data.replace("delete_stanok_obrabotka_", "")
        if obrabotka not in data["obrabotka"]:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Обработка топилмади!"
            )
            return
        keyboard = [[InlineKeyboardButton(stanok, callback_data=f"delete_stanok_{obrabotka}_{stanok}")] for stanok in sorted(data["obrabotka"][obrabotka])]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.user_data["last_message_id"] = (await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Ўчириладиган станокни танланг:",
            reply_markup=reply_markup
        )).message_id
        return

    if query.data.startswith("delete_stanok_"):
        _, obrabotka, stanok = query.data.split("_", 2)
        if obrabotka in data["obrabotka"] and stanok in data["obrabotka"][obrabotka]:
            data["obrabotka"][obrabotka].remove(stanok)
            save_data(data)
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text=f"Станок '{stanok}' ўчирилди!"
            )
        else:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Станок топилмади!"
            )
        return

    if query.data.startswith("delete_qalinlik_"):
        qalinlik = query.data.replace("delete_qalinlik_", "")
        if qalinlik in data["qalinlik"]:
            data["qalinlik"].remove(qalinlik)
            save_data(data)
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text=f"Қалинлик '{qalinlik}' ўчирилди!"
            )
        else:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Қалинлик топилмади!"
            )
        return

    if query.data.startswith("delete_mahsulot_katalog_"):
        katalog_name = query.data.replace("delete_mahsulot_katalog_", "")
        if katalog_name not in data["katalog"]:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Каталог топилмади!"
            )
            return
        keyboard = [[InlineKeyboardButton(mahsulot["nom"], callback_data=f"delete_mahsulot_{katalog_name}_{mahsulot['id']}")] for mahsulot in sorted(data["katalog"][katalog_name]["items"], key=lambda x: x["nom"])]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.user_data["last_message_id"] = (await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Ўчириладиган тошни танланг:",
            reply_markup=reply_markup
        )).message_id
        return

    if query.data.startswith("delete_mahsulot_"):
        _, katalog_name, mahsulot_id = query.data.split("_", 2)
        if katalog_name not in data["katalog"]:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Каталог топилмади!"
            )
            return
        mahsulot = next((item for item in data["katalog"][katalog_name]["items"] if item["id"] == mahsulot_id), None)
        if not mahsulot:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text="Тош топилмади!"
            )
            return
        if mahsulot.get("rasm_path") and os.path.exists(mahsulot["rasm_path"]):
            os.remove(mahsulot["rasm_path"])
        data["katalog"][katalog_name]["items"] = [item for item in data["katalog"][katalog_name]["items"] if item["id"] != mahsulot_id]
        save_data(data)
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=f"Тош '{mahsulot['nom']}' ўчирилди!"
        )
        return

def main():
    application = Application.builder().token(BOT_TOKEN).build()
    
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("rasm", rasm))
    application.add_handler(CommandHandler("rasmsiz", rasmsiz))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    application.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    application.add_handler(CallbackQueryHandler(button))
    
    application.job_queue.run_daily(
        send_daily_report,
        time(hour=23, minute=59, tzinfo=TASHKENT_TZ)
    )
    application.job_queue.run_daily(
        send_10day_report,
        time(hour=23, minute=59, tzinfo=TASHKENT_TZ)
    )
    
    application.run_polling()

if __name__ == "__main__":
    main()